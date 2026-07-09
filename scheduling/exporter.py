"""Sprint 2.2 Final - Exporters"""
from abc import ABC, abstractmethod
import csv, json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font

class ScheduleExporter(ABC):
    @abstractmethod
    def export(self, schedule, output_path): ...

class _Helper:
    @staticmethod
    def metadata(s):
        return [("Schedule ID",s.schedule_id),("Academic Year",s.academic_year),
                ("Division",s.division),("Level",s.level),("Term",s.term),
                ("Version",s.version),("Exported On",datetime.now().strftime("%Y-%m-%d %H:%M:%S"))]
    @staticmethod
    def headers():
        return ["Cycle Day","Slot No","Event ID","Parent Event","Event Type","Event Name","Start Time","End Time","Activity","Notes"]
    @staticmethod
    def rows(s):
        for e in sorted(s.entries,key=lambda x:(x.cycle_day,x.start_time,x.end_time)):
            yield [e.cycle_day,e.slot_no,e.event_id,e.parent_event,e.event_type,e.event_name,e.start_time,e.end_time,e.activity,e.notes]

class JSONExporter(ScheduleExporter):
    def export(self,s,o):
        p=Path(o); p.parent.mkdir(parents=True,exist_ok=True)
        with p.open("w",encoding="utf-8") as f: json.dump(s.to_dict(),f,indent=4,ensure_ascii=False)

class CSVExporter(ScheduleExporter):
    def export(self,s,o):
        p=Path(o); p.parent.mkdir(parents=True,exist_ok=True)
        with p.open("w",newline="",encoding="utf-8") as f:
            w=csv.writer(f)
            for r in _Helper.metadata(s): w.writerow(r)
            w.writerow([])
            w.writerow(_Helper.headers())
            for r in _Helper.rows(s): w.writerow(r)

class ExcelExporter(ScheduleExporter):
    def export(self,s,o):
        p=Path(o); p.parent.mkdir(parents=True,exist_ok=True)
        wb=Workbook(); ws=wb.active; ws.title="Schedule"
        ws["A1"]="SGS Academic Planner"; ws["A1"].font=Font(bold=True,size=16)
        ws["A2"]="Schedule Export"; ws["A2"].font=Font(bold=True)
        r=4
        for k,v in _Helper.metadata(s):
            ws.cell(r,1).value=k; ws.cell(r,1).font=Font(bold=True); ws.cell(r,2).value=v; r+=1
        r+=1
        for c,h in enumerate(_Helper.headers(),1):
            ws.cell(r,c).value=h; ws.cell(r,c).font=Font(bold=True)
        for row in _Helper.rows(s):
            r+=1
            for c,v in enumerate(row,1): ws.cell(r,c).value=v
        ws.freeze_panes="A12"
        for col in ws.columns:
            l=max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width=min(l+3,40)
        wb.save(p)
