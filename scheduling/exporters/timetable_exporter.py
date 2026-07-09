"""
scheduling/exporters/timetable_exporter.py

Operational timetable exporter.

Creates a day x time-slot view of a Schedule.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


class TimetableExporter:

    def export(self, schedule, output_file: str):

        wb = Workbook()
        ws = wb.active
        ws.title = "Timetable"

        # Build ordered list of unique time slots
        time_slots = sorted(
            {(e.start_time, e.end_time) for e in schedule.entries},
            key=lambda x: x[0]
        )

        # Header
        ws.cell(row=1, column=1).value = "Day"

        for col, (start, end) in enumerate(time_slots, start=2):
            c = ws.cell(row=1, column=col)
            c.value = f"{start}-{end}"
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center")

        # Fast lookup
        lookup = {
            (e.cycle_day, e.start_time, e.end_time): e.activity
            for e in schedule.entries
        }

        grey = PatternFill(fill_type="solid", fgColor="F2F2F2")
        white = PatternFill(fill_type="solid", fgColor="FFFFFF")
        thin = Side(style="thin")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        row = 2

        for day in range(1, schedule.total_days() + 1):

            ws.cell(row=row, column=1).value = f"Day {day}"
            ws.cell(row=row, column=1).font = Font(bold=True)

            fill = grey if day % 2 == 0 else white

            for col, (start, end) in enumerate(time_slots, start=2):
                cell = ws.cell(row=row, column=col)
                cell.value = lookup.get((day, start, end), "")
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center")

            row += 1

        # Freeze header and first column
        ws.freeze_panes = "B2"

        # Auto width
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            width = 14
            for cell in ws[letter]:
                if cell.value:
                    width = max(width, len(str(cell.value)) + 2)
            ws.column_dimensions[letter].width = min(width, 30)

        wb.save(output_file)
